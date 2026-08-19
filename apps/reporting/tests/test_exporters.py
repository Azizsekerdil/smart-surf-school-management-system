"""The export engines. These run without a database on purpose."""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal

import pytest
from django.utils import timezone
from openpyxl import load_workbook

from apps.reporting.exporters import get_exporter
from apps.reporting.exporters.base import ColumnKind, ReportData, display_value
from apps.reporting.exporters.csv_export import ENCODING
from apps.reporting.exporters.pdf import resolve_fonts
from apps.reporting.exporters.registry import UnknownExportFormat

#: A Turkish row, because the whole point of the font work is that this survives.
TURKISH_NAME = "Şükrü Çağrı Öztürk"


def sample_report(rows: int = 3, columns: int = 4) -> ReportData:
    all_columns = [
        ("Kod", ColumnKind.TEXT),
        ("Müşteri", ColumnKind.TEXT),
        ("Tutar", ColumnKind.MONEY),
        ("Tarih", ColumnKind.DATE),
        ("Oran", ColumnKind.PERCENT),
        ("Adet", ColumnKind.NUMBER),
        ("Aktif", ColumnKind.BOOLEAN),
        ("Not", ColumnKind.TEXT),
    ][:columns]

    return ReportData(
        title="Gelir Raporu",
        subtitle="Son 30 gün",
        columns=[name for name, _kind in all_columns],
        column_kinds=[kind for _name, kind in all_columns],
        rows=[
            [
                f"INV-{index:03d}",
                TURKISH_NAME,
                Decimal("1250.50"),
                date(2026, 8, 18),
                Decimal("12.5"),
                7,
                True,
                "Uzun bir açıklama " * 4,
            ][:columns]
            for index in range(rows)
        ],
        summary={"Toplam": (Decimal("3751.50"), ColumnKind.MONEY), "Satır": 3},
        filters={"Dönem": "Son 30 gün", "Durum": "Ödendi"},
        currency="₺",
    )


# ---------------------------------------------------------------------------
# Registry
# ---------------------------------------------------------------------------
@pytest.mark.parametrize("fmt", ["pdf", "excel", "csv", "PDF", " Excel "])
def test_registry_resolves_known_formats(fmt):
    assert get_exporter(fmt) is not None


def test_registry_rejects_an_unknown_format():
    with pytest.raises(UnknownExportFormat):
        get_exporter("docx")


# ---------------------------------------------------------------------------
# Formatting
# ---------------------------------------------------------------------------
def test_money_is_formatted_with_the_currency_symbol():
    assert "₺" in display_value(Decimal("10.00"), ColumnKind.MONEY, "₺")


def test_none_renders_as_an_empty_cell_not_the_word_none():
    assert display_value(None) == ""


def test_booleans_never_leak_python_repr():
    assert display_value(True, ColumnKind.BOOLEAN) not in {"True", "1"}


def test_duration_reads_like_a_lesson_board():
    assert display_value(150, ColumnKind.DURATION) == "2h 30m"


def test_wide_tables_switch_to_landscape():
    assert sample_report(columns=8).effective_orientation == "landscape"
    assert sample_report(columns=4).effective_orientation == "portrait"


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------
def test_pdf_font_registration_prefers_a_unicode_font():
    regular, bold = resolve_fonts()
    assert regular and bold
    # Helvetica has no Turkish glyphs; the bundled Vera face must win.
    assert regular != "Helvetica"


def test_pdf_renders_a_real_document():
    payload = get_exporter("pdf").render(sample_report(rows=5))
    assert payload.startswith(b"%PDF")
    assert payload.endswith(b"%%EOF\n") or b"%%EOF" in payload[-32:]


def test_pdf_handles_an_empty_report_without_crashing():
    data = ReportData(title="Boş", columns=["A", "B"], rows=[], message="No rows.")
    payload = get_exporter("pdf").render(data)
    assert payload.startswith(b"%PDF")


def test_pdf_paginates_a_long_table():
    """Page N of M needs the two-pass canvas; a long table proves it runs."""
    long_report = sample_report(rows=400)
    payload = get_exporter("pdf").render(long_report)
    assert payload.startswith(b"%PDF")
    assert payload.count(b"/Type /Page") > 1 or payload.count(b"/Type/Page") > 1


def test_pdf_survives_a_very_wide_table():
    payload = get_exporter("pdf").render(sample_report(rows=3, columns=8))
    assert payload.startswith(b"%PDF")


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def test_excel_writes_typed_cells_and_a_summary_sheet():
    payload = get_exporter("excel").render(sample_report(rows=3))
    workbook = load_workbook(io.BytesIO(payload))

    assert len(workbook.sheetnames) == 2
    data_sheet = workbook[workbook.sheetnames[0]]

    # Header styling and a frozen pane so the header survives scrolling.
    assert data_sheet["A1"].font.bold is True
    assert data_sheet.freeze_panes == "A2"
    assert data_sheet.auto_filter.ref is not None

    # Money must arrive as a number, not as text, or the workbook is useless.
    assert isinstance(data_sheet["C2"].value, float)
    assert "#,##0.00" in data_sheet["C2"].number_format

    summary_sheet = workbook[workbook.sheetnames[1]]
    assert summary_sheet["A1"].value == "Gelir Raporu"


def test_excel_column_widths_are_set():
    payload = get_exporter("excel").render(sample_report())
    workbook = load_workbook(io.BytesIO(payload))
    sheet = workbook[workbook.sheetnames[0]]
    assert sheet.column_dimensions["A"].width >= 10


def test_excel_writes_naive_datetimes():
    """openpyxl refuses timezone-aware datetimes; the exporter must convert."""
    data = ReportData(
        title="Zaman",
        columns=["When"],
        column_kinds=[ColumnKind.DATETIME],
        rows=[[timezone.now()]],
    )
    workbook = load_workbook(io.BytesIO(get_exporter("excel").render(data)))
    cell = workbook[workbook.sheetnames[0]]["A2"]
    assert isinstance(cell.value, datetime)
    assert cell.value.tzinfo is None


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------
def test_csv_starts_with_a_bom_so_turkish_excel_reads_utf8():
    payload = get_exporter("csv").render(sample_report())
    assert payload.startswith(b"\xef\xbb\xbf")


def test_csv_uses_the_semicolon_delimiter():
    payload = get_exporter("csv").render(sample_report())
    first_line = payload.decode(ENCODING).splitlines()[0]
    assert first_line.count(";") == 3


def test_csv_keeps_turkish_characters_intact():
    text = get_exporter("csv").render(sample_report()).decode(ENCODING)
    assert TURKISH_NAME in text


def test_csv_appends_the_summary_after_a_blank_line():
    text = get_exporter("csv").render(sample_report()).decode(ENCODING)
    body, _blank, tail = text.partition("\r\n\r\n")
    assert body.count("\r\n") == 3  # header + 3 rows
    assert "Toplam" in tail


def test_csv_explains_an_empty_result():
    data = ReportData(title="Boş", columns=["A"], rows=[], message="Nothing matched.")
    text = get_exporter("csv").render(data).decode(ENCODING)
    assert "Nothing matched." in text


# ---------------------------------------------------------------------------
# Filenames
# ---------------------------------------------------------------------------
def test_filenames_are_ascii_and_dated():
    data = sample_report()
    for fmt, extension in (("pdf", "pdf"), ("excel", "xlsx"), ("csv", "csv")):
        name = get_exporter(fmt).filename(data)
        assert name.isascii()
        assert name.endswith(f".{extension}")
        assert timezone.localdate().strftime("%Y-%m-%d") in name
