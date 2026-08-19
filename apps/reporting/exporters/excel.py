"""Excel export via openpyxl.

Two sheets on purpose
---------------------
``Data`` stays a clean rectangle — header row, then rows, nothing else — so it
can be filtered, pivoted or fed straight into an accountant's workbook. All the
provenance (title, period, filters, totals) lives on ``Summary``. A merged title
banner on top of the data would break the autofilter and every formula written
against the sheet.

Values are written as real types: ``Decimal`` amounts become numbers with a
currency format, dates become dates. A money column exported as text is useless
in a spreadsheet, which is the whole point of offering this format.
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime, time
from decimal import Decimal

from django.utils import timezone
from django.utils.translation import gettext as _
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .base import BaseExporter, ColumnKind, ReportData, display_value

BRAND_FILL = PatternFill("solid", fgColor="0083CE")
PANEL_FILL = PatternFill("solid", fgColor="F1F5F9")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
LABEL_FONT = Font(bold=True, color="0F172A", size=11)
TITLE_FONT = Font(bold=True, color="075985", size=15)
THIN = Side(style="thin", color="CBD5E1")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DATE_FORMAT = "DD.MM.YYYY"
DATETIME_FORMAT = "DD.MM.YYYY HH:MM"
NUMBER_FORMAT = "#,##0.##"
PERCENT_FORMAT = '#,##0.0"%"'

MIN_COLUMN_WIDTH = 10
MAX_COLUMN_WIDTH = 52
#: Rows sampled when sizing columns; scanning a 20 000-row export twice is waste.
WIDTH_SAMPLE_ROWS = 400

#: Characters Excel forbids in a sheet name.
_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


class ExcelExporter(BaseExporter):
    """XLSX workbook with a filterable data sheet and a summary sheet."""

    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    file_extension = "xlsx"
    label = "Excel"

    def render(self, data: ReportData) -> bytes:
        workbook = Workbook()
        data_sheet = workbook.active
        data_sheet.title = _sheet_name(_("Data"))
        self._write_data(data_sheet, data)

        summary_sheet = workbook.create_sheet(_sheet_name(_("Summary")))
        self._write_summary(summary_sheet, data)

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    # --- data sheet -------------------------------------------------------
    def _write_data(self, sheet: Worksheet, data: ReportData) -> None:
        money_format = _money_format(data.currency)

        if not data.columns:
            sheet["A1"] = str(data.message or _("This report produced no columns."))
            sheet.column_dimensions["A"].width = 80
            return

        for index, header in enumerate(data.columns, start=1):
            cell = sheet.cell(row=1, column=index, value=str(header))
            cell.font = HEADER_FONT
            cell.fill = BRAND_FILL
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = CELL_BORDER
        sheet.row_dimensions[1].height = 22

        for row_index, row in enumerate(data.rows, start=2):
            for column_index, value in enumerate(row, start=1):
                kind = data.kind(column_index - 1)
                cell = sheet.cell(row=row_index, column=column_index)
                cell.value = _cell_value(value, kind)
                number_format = _number_format_for(kind, money_format)
                if number_format:
                    cell.number_format = number_format
                    cell.alignment = Alignment(horizontal="right")
                cell.border = CELL_BORDER

        # Freeze the header and let the user filter without extra clicks.
        sheet.freeze_panes = "A2"
        last_column = get_column_letter(len(data.columns))
        last_row = max(len(data.rows) + 1, 1)
        sheet.auto_filter.ref = f"A1:{last_column}{last_row}"

        self._autosize(sheet, data)

    def _autosize(self, sheet: Worksheet, data: ReportData) -> None:
        sample = data.rows[:WIDTH_SAMPLE_ROWS]
        for index, header in enumerate(data.columns):
            longest = len(str(header))
            kind = data.kind(index)
            for row in sample:
                if index >= len(row):
                    continue
                longest = max(longest, len(display_value(row[index], kind, data.currency)))
            width = min(max(longest + 3, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH)
            sheet.column_dimensions[get_column_letter(index + 1)].width = width

    # --- summary sheet ----------------------------------------------------
    def _write_summary(self, sheet: Worksheet, data: ReportData) -> None:
        sheet.column_dimensions["A"].width = 34
        sheet.column_dimensions["B"].width = 46

        row = 1
        sheet.cell(row=row, column=1, value=str(data.title)).font = TITLE_FONT
        row += 1
        if data.subtitle:
            sheet.cell(row=row, column=1, value=str(data.subtitle))
            row += 1

        row += 1
        row = self._write_pairs(
            sheet,
            row,
            _("Report details"),
            [
                (_("Generated at"), display_value(data.generated_at)),
                (_("Rows"), display_value(data.row_count, ColumnKind.NUMBER)),
            ]
            + ([(_("Row limit reached"), display_value(data.truncated_at, ColumnKind.NUMBER))]
               if data.truncated_at else []),
        )

        filters = data.filter_items()
        if filters:
            row += 1
            row = self._write_pairs(sheet, row, _("Applied filters"), filters)

        summary = data.summary_items()
        if summary:
            row += 1
            row = self._write_pairs(sheet, row, _("Summary"), summary)

        if data.message:
            row += 1
            sheet.cell(row=row, column=1, value=str(data.message)).alignment = Alignment(
                wrap_text=True
            )

    def _write_pairs(self, sheet: Worksheet, row: int, heading: str, pairs: list) -> int:
        cell = sheet.cell(row=row, column=1, value=str(heading))
        cell.font = LABEL_FONT
        cell.fill = PANEL_FILL
        sheet.cell(row=row, column=2).fill = PANEL_FILL
        row += 1
        for label, value in pairs:
            sheet.cell(row=row, column=1, value=str(label)).font = Font(bold=True)
            sheet.cell(row=row, column=2, value=str(value))
            row += 1
        return row


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _sheet_name(value: str) -> str:
    """Excel sheet names: at most 31 characters and no ``[]:*?/\\``."""
    cleaned = _INVALID_SHEET_CHARS.sub("-", str(value)).strip() or "Sheet"
    return cleaned[:31]


def _money_format(symbol: str) -> str:
    if not symbol:
        return "#,##0.00"
    escaped = symbol.replace('"', "")
    return f'#,##0.00 "{escaped}"'


def _number_format_for(kind: str, money_format: str) -> str:
    if kind == ColumnKind.MONEY:
        return money_format
    if kind == ColumnKind.PERCENT:
        return PERCENT_FORMAT
    if kind == ColumnKind.NUMBER:
        return NUMBER_FORMAT
    if kind == ColumnKind.DATE:
        return DATE_FORMAT
    if kind == ColumnKind.DATETIME:
        return DATETIME_FORMAT
    return ""


def _cell_value(value, kind: str):
    """Return something openpyxl can store natively, else readable text."""
    if value is None:
        return None
    if isinstance(value, bool):
        return _("Yes") if value else _("No")
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, datetime):
        # openpyxl cannot write timezone-aware datetimes.
        return timezone.localtime(value).replace(tzinfo=None) if timezone.is_aware(value) else value
    if isinstance(value, (date, time)):
        return value
    if kind == ColumnKind.DURATION:
        return display_value(value, kind)
    return str(value)


__all__ = ["ExcelExporter"]
